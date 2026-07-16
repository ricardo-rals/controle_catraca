import csv
from io import BytesIO

import pandas as pd
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from weasyprint import HTML

from .registry import RELATORIOS, get_relatorio

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PREVIEW_MAX = 50  # linhas mostradas na prévia (a exportação leva tudo)


@login_required
def lista(request):
    """GET /relatorios/ — central com os relatórios disponíveis (HU-040)."""
    return render(request, "relatorios/lista.html", {"relatorios": RELATORIOS.values()})


@login_required
def detalhe(request, slug):
    """GET /relatorios/<slug>/ — filtros + prévia + export.

    Não busca dados ao abrir: só monta o relatório quando o usuário envia a
    busca (parâmetro `buscar` — vem do "Filtrar" ou do "Buscar todos os dados").
    """
    rel = get_relatorio(slug)
    contexto = {
        "rel": rel,
        "form": rel.form_builder(request),
        "buscou": "buscar" in request.GET,
        "dados": None,
    }
    if contexto["buscou"]:
        dados = rel.montar(request)
        dados["preview"] = dados["linhas"][:PREVIEW_MAX]
        contexto["dados"] = dados
    return render(request, "relatorios/detalhe.html", contexto)


@login_required
def exportar(request, slug, formato):
    """GET /relatorios/<slug>/export/<formato>/ — gera o arquivo com os filtros."""
    rel = get_relatorio(slug)
    dados = rel.montar(request)
    nome = f"{slug}_{timezone.localdate():%Y%m%d}"

    if formato == "csv":
        return _exportar_csv(dados, nome)
    if formato == "excel":
        return _exportar_excel(dados, nome)
    if formato == "pdf":
        return _exportar_pdf(request, rel, dados, nome)
    raise Http404("Formato inválido (use pdf, excel ou csv).")


def _exportar_csv(dados, nome):
    resposta = HttpResponse(content_type="text/csv; charset=utf-8")
    resposta["Content-Disposition"] = f'attachment; filename="{nome}.csv"'
    resposta.write("﻿")  # BOM: Excel abre acentuação corretamente
    escritor = csv.writer(resposta)
    escritor.writerow(dados["colunas"])
    escritor.writerows(dados["linhas"])
    return resposta


def _exportar_excel(dados, nome):
    resumo_df = pd.DataFrame(dados["resumo"], columns=["Indicador", "Valor"])
    dados_df = pd.DataFrame(dados["linhas"] or None, columns=dados["colunas"])

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        resumo_df.to_excel(writer, sheet_name="Resumo", index=False)
        dados_df.to_excel(writer, sheet_name="Dados", index=False)
        _formatar_planilha(writer.sheets["Resumo"])
        _formatar_planilha(writer.sheets["Dados"])

    resposta = HttpResponse(buffer.getvalue(), content_type=XLSX_CONTENT_TYPE)
    resposta["Content-Disposition"] = f'attachment; filename="{nome}.xlsx"'
    return resposta


def _formatar_planilha(ws):
    """Cabeçalho em negrito + largura automática das colunas."""
    for coluna in ws.columns:
        largura = max(
            (len(str(c.value)) for c in coluna if c.value is not None), default=0
        )
        ws.column_dimensions[get_column_letter(coluna[0].column)].width = min(
            largura + 2, 60
        )
    for celula in ws[1]:
        celula.font = Font(bold=True)


def _exportar_pdf(request, rel, dados, nome):
    html = render_to_string(
        "relatorios/relatorio_pdf.html",
        {
            "rel": rel,
            "dados": dados,
            "gerado_em": timezone.localtime(),
            "usuario": request.user,
        },
    )
    pdf = HTML(string=html, base_url=request.build_absolute_uri("/")).write_pdf()
    resposta = HttpResponse(pdf, content_type="application/pdf")
    resposta["Content-Disposition"] = f'attachment; filename="{nome}.pdf"'
    return resposta
