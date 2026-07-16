import logging

import pandas as pd
from django.db import transaction
from django.db.models import Q

from apps.acessos.models import PontoAcesso, RegistroAcesso
from .utils.pseudonimizacao import criptografar_valor
from .models import Importacao, FalhaImportacao

logger = logging.getLogger(__name__)

TAMANHO_LOTE = 1000


def _texto_opcional(valor):
    """Converte NaN/vazio do pandas para None; caso contrário, string trim."""
    if valor is None or pd.isna(valor):
        return None
    texto = str(valor).strip()
    return texto or None


def _extrair_hyperlinks_coluna(caminho_xlsx, nome_coluna):
    """Lê apenas os hyperlinks de uma coluna do xlsx (por nome no header).

    Retorna {indice_df: url}. Índice do df = linha_planilha - 2 (cabeçalho + 1-based).
    Célula sem hyperlink não entra no dict — o valor original do pandas se mantém.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(caminho_xlsx, data_only=True, read_only=False)
    except Exception:
        return {}
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if nome_coluna not in header:
        return {}
    col_idx = header.index(nome_coluna) + 1
    urls = {}
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        link = getattr(cell, "hyperlink", None)
        alvo = getattr(link, "target", None) if link else None
        if alvo:
            urls[row_idx - 2] = alvo
    return urls


class ImportacaoService:
    """
    Responsável pelo pipeline de dados (Pandas) e persistência em lote.

    Critérios de aceite (HU-019 e HU-021):
      - Carregar CSV com Pandas
      - Validar e descartar dados inválidos
      - Transformação (LGPD)
      - Mapeamento de chaves estrangeiras (PontoAcesso)
      - Desduplicação em memória (Pandas)
      - Desduplicação contra o banco de dados (Query Otimizada)
      - Inserção via bulk_create em lotes de 1000
    """

    def __init__(self, arquivo, importacao: Importacao):
        self.arquivo = arquivo
        self.importacao = importacao
        self._falhas: list[FalhaImportacao] = []

    def processar(self):
        """
        Ponto de entrada: processa o arquivo CSV e persiste os registros válidos.

        Não re-levanta exceções: em caso de erro a importação fica com status
        FALHA e o motivo preenchido, e a view decide o feedback pelo status.
        """
        try:
            with transaction.atomic():
                self._executar_pipeline()
        except Exception as exc:
            logger.exception(
                "Erro inesperado ao processar importação %s", self.importacao.pk
            )
            self._marcar_como_falha(exc)

        return self.importacao

    # Cabeçalho real do export da catraca (Title-Case com espaços).
    COL_CREDENCIAL = "Número da Credencial"
    COL_NOME = "Nome"
    COL_DATA = "Data do Evento"
    COL_EQUIPAMENTO = "Equipamento"
    COL_DIRECAO = "Direção do Evento"
    COL_GRUPO = "Grupo de Equipamento"
    COL_AREA_ORIGEM = "Área de Origem"
    COL_AREA_DESTINO = "Área de Destino"
    COL_EVENTO = "Evento"
    COL_FOTO = "Foto"
    COL_TIPO_CONSULTA = "Tipo de Consulta"
    COLUNAS_OBRIGATORIAS = (COL_CREDENCIAL, COL_DATA)

    def _executar_pipeline(self):
        # 1. Carrega CSV ou Excel conforme a extensão (tudo como texto p/ validar).
        eh_xlsx = isinstance(self.arquivo, str) and self.arquivo.lower().endswith(
            (".xlsx", ".xls")
        )
        if eh_xlsx:
            df = pd.read_excel(self.arquivo, dtype=str)
        else:
            df = pd.read_csv(self.arquivo, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        # A coluna Foto no xlsx mostra "Ver Imagem" mas o valor real é o hyperlink
        # da célula (URL do ImageHandler). read_excel só traz o texto visível.
        if eh_xlsx and self.COL_FOTO in df.columns:
            urls_foto = _extrair_hyperlinks_coluna(self.arquivo, self.COL_FOTO)
            if urls_foto:
                df[self.COL_FOTO] = (
                    df.index.to_series().map(urls_foto).fillna(df[self.COL_FOTO])
                )

        # HU-018: cabeçalho obrigatório. Coluna essencial ausente aborta com erro claro.
        faltando = [c for c in self.COLUNAS_OBRIGATORIAS if c not in df.columns]
        if faltando:
            raise ValueError(
                "Cabeçalho inválido: coluna(s) obrigatória(s) ausente(s): "
                + ", ".join(faltando)
            )

        # Normaliza para os nomes internos usados pelo restante do pipeline.
        df["credencial"] = df[self.COL_CREDENCIAL]
        df["nome"] = df.get(self.COL_NOME, "")
        df["timestamp"] = df[self.COL_DATA]
        df["ponto_acesso_nome"] = df.get(self.COL_EQUIPAMENTO)
        df["status_acesso"] = df.get(self.COL_DIRECAO, "")
        df["grupo_equipamento"] = df.get(self.COL_GRUPO)
        df["area_origem"] = df.get(self.COL_AREA_ORIGEM)
        df["area_destino"] = df.get(self.COL_AREA_DESTINO)
        df["evento"] = df.get(self.COL_EVENTO)
        df["foto"] = df.get(self.COL_FOTO)
        df["tipo_consulta"] = df.get(self.COL_TIPO_CONSULTA)

        # 2. Inválidos: registra (linha + motivo) linhas com credencial ou data vazias.
        falta_cred = df["credencial"].isna() | df["credencial"].astype(
            str
        ).str.strip().isin(["", "nan"])
        falta_data = df["timestamp"].isna() | df["timestamp"].astype(
            str
        ).str.strip().isin(["", "nan"])
        for idx in df.index[falta_cred | falta_data]:
            motivos = []
            if falta_cred[idx]:
                motivos.append("credencial vazia")
            if falta_data[idx]:
                motivos.append("data do evento vazia")
            self._registrar_falha(idx, " | ".join(motivos))
        df = df[~(falta_cred | falta_data)]

        # Converte data BR (dd/mm/aaaa hh:mm:ss). Data não parseável vira falha.
        df["timestamp_dt"] = pd.to_datetime(
            df["timestamp"], errors="coerce", dayfirst=True, utc=True
        )
        data_invalida = df["timestamp_dt"].isna()
        for idx in df.index[data_invalida]:
            self._registrar_falha(idx, "data inválida")
        df = df[~data_invalida]

        self.importacao.total_invalidos = len(self._falhas)

        if df.empty:
            self._finalizar_importacao(df)
            return

        df["credencial_cifrada"] = df["credencial"].apply(
            lambda x: criptografar_valor(str(x).strip(), deterministico=True)
        )
        df["nome_cifrado"] = df["nome"].apply(
            lambda x: criptografar_valor(str(x).strip()) if _texto_opcional(x) else ""
        )

        # 4. Transformação 2 (FK): Mapeie os nomes dos equipamentos (PontoAcesso).
        # Cada equipamento carrega seu grupo (ex.: "PORTARIA"), usado tanto no
        # campo grupo_equipamento quanto como localização.
        nomes_pontos = df["ponto_acesso_nome"].dropna().unique()
        pontos_db = PontoAcesso.objects.filter(nome__in=nomes_pontos)
        mapa_pontos = {p.nome: p.id for p in pontos_db}

        grupo_por_equipamento = (
            df.dropna(subset=["ponto_acesso_nome"])
            .groupby("ponto_acesso_nome")["grupo_equipamento"]
            .apply(
                lambda s: next((v for v in s if pd.notna(v) and str(v).strip()), None)
            )
            .to_dict()
        )

        # Preenche o grupo em equipamentos criados por importações antigas,
        # que nasceram sem essa informação (reimportar o xlsx corrige o dado).
        pontos_para_atualizar = []
        for ponto in pontos_db:
            grupo = grupo_por_equipamento.get(ponto.nome)
            if grupo and not ponto.grupo_equipamento:
                ponto.grupo_equipamento = grupo
                if ponto.localizacao == ponto.nome:
                    ponto.localizacao = grupo
                pontos_para_atualizar.append(ponto)
        if pontos_para_atualizar:
            PontoAcesso.objects.bulk_update(
                pontos_para_atualizar, ["grupo_equipamento", "localizacao"]
            )

        pontos_faltantes = set(nomes_pontos) - set(mapa_pontos.keys())
        for nome_ponto in pontos_faltantes:
            if str(nome_ponto).strip():
                grupo = grupo_por_equipamento.get(nome_ponto)
                novo_ponto = PontoAcesso.objects.create(
                    nome=nome_ponto,
                    localizacao=grupo or nome_ponto,
                    grupo_equipamento=grupo,
                )
                mapa_pontos[nome_ponto] = novo_ponto.id

        df["ponto_acesso_id"] = df["ponto_acesso_nome"].map(mapa_pontos)

        # 5. Duplicados Internos: Utilize drop_duplicates do Pandas
        antes_dedup = len(df)
        df.drop_duplicates(
            subset=["credencial_cifrada", "timestamp_dt", "ponto_acesso_id"],
            inplace=True,
        )
        self.importacao.total_duplicados = antes_dedup - len(df)

        if df.empty:
            self._finalizar_importacao(df)
            return

        # 6. Confronto com o Banco de Dados.
        # Para cada linha do df que bate com um registro existente no banco:
        #   - se o registro do banco tem campos vazios que a nova linha preenche,
        #     enriquecemos o registro (bulk_update) → conta em total_atualizados
        #   - senão, é duplicata verdadeira → conta em total_duplicados
        # Só linhas que NÃO estão no banco vão para bulk_create.
        existentes_map = self._buscar_existentes(df)

        if existentes_map:
            df, enriquecidos, duplicatas_puras = self._separar_novos_e_enriquecidos(
                df, existentes_map
            )
            if enriquecidos:
                RegistroAcesso.objects.bulk_update(
                    enriquecidos,
                    fields=[
                        "credencial_cifrada",
                        "nome_cifrado",
                        "tipo_acesso",
                        "evento",
                        "foto",
                        "tipo_consulta",
                        "area_origem",
                        "area_destino",
                    ],
                    batch_size=TAMANHO_LOTE,
                )
            self.importacao.total_atualizados = len(enriquecidos)
            self.importacao.total_duplicados += duplicatas_puras

        # 7. Finalização
        # Contabilize o que sobrou no DataFrame como total_validos.
        self._finalizar_importacao(df)
        self._inserir_no_banco(df)

    def _buscar_existentes(self, df):
        """Retorna dict {(cred_cifrada, timestamp_utc, ponto_id): RegistroAcesso} do banco
        para as tuplas do df, em lotes para não estourar o tamanho do WHERE."""
        existentes = {}
        lote_size = 900
        for i in range(0, len(df), lote_size):
            lote = df.iloc[i : i + lote_size]
            q_objects = Q()
            for row in lote[
                ["credencial_cifrada", "timestamp_dt", "ponto_acesso_id"]
            ].itertuples(index=False):
                if pd.notna(row[2]):
                    q_objects |= Q(
                        credencial_cifrada=row[0],
                        timestamp=row[1],
                        ponto_acesso_id=row[2],
                    )
                else:
                    q_objects |= Q(
                        credencial_cifrada=row[0],
                        timestamp=row[1],
                        ponto_acesso__isnull=True,
                    )
            if not q_objects:
                continue
            for obj in RegistroAcesso.objects.filter(q_objects):
                chave = (
                    obj.credencial_cifrada,
                    pd.Timestamp(obj.timestamp),
                    obj.ponto_acesso_id,
                )
                existentes[chave] = obj
        return existentes

    _CAMPOS_ENRIQUECIVEIS = (
        ("credencial_cifrada", "credencial_cifrada"),
        ("nome_cifrado", "nome_cifrado"),
        ("status_acesso", "tipo_acesso"),
        ("evento", "evento"),
        ("foto", "foto"),
        ("tipo_consulta", "tipo_consulta"),
        ("area_origem", "area_origem"),
        ("area_destino", "area_destino"),
    )

    def _separar_novos_e_enriquecidos(self, df, existentes_map):
        """Divide o df em (novos, enriquecidos, duplicatas_puras).

        - novos: linhas que não existem no banco (viram bulk_create)
        - enriquecidos: objetos do banco que ganham campos antes vazios
        - duplicatas_puras: registros já completos, apenas contados
        """
        novos_idx = []
        enriquecidos = []
        duplicatas_puras = 0

        for idx, row in df.iterrows():
            ponto_id = None
            if pd.notna(row["ponto_acesso_id"]):
                ponto_id = int(row["ponto_acesso_id"])
            chave = (
                row["credencial_cifrada"],
                row["timestamp_dt"],
                ponto_id,
            )
            existente = existentes_map.get(chave)
            if existente is None:
                novos_idx.append(idx)
                continue

            alterado = False
            for campo_df, campo_model in self._CAMPOS_ENRIQUECIVEIS:
                atual = getattr(existente, campo_model, None)
                novo = _texto_opcional(row.get(campo_df))
                if self._deve_enriquecer(campo_model, atual, novo):
                    setattr(existente, campo_model, novo)
                    alterado = True
            if alterado:
                enriquecidos.append(existente)
            else:
                duplicatas_puras += 1

        return df.loc[novos_idx], enriquecidos, duplicatas_puras

    @staticmethod
    def _deve_enriquecer(campo, atual, novo):
        """Só enriquece se o novo valor tem conteúdo e o atual está vazio.

        Trata 'Ver Imagem' no campo foto como placeholder (o xlsx exporta esse
        texto quando o valor real deveria ser o URL do hyperlink).
        """
        if not novo:
            return False
        if not atual:
            return True
        if (
            campo == "foto"
            and str(atual).strip() == "Ver Imagem"
            and novo != "Ver Imagem"
        ):
            return True
        return False

    def _registrar_falha(self, idx, motivo):
        # idx: índice 0-based do read_csv → linha no arquivo = idx + 2 (cabeçalho + base 1)
        self._falhas.append(
            FalhaImportacao(
                importacao=self.importacao,
                linha_arquivo=int(idx) + 2,
                motivo_erro=motivo,
            )
        )

    def _finalizar_importacao(self, df):
        self.importacao.total_validos = len(df)
        self.importacao.total_registros = (
            len(df)
            + self.importacao.total_invalidos
            + self.importacao.total_duplicados
            + self.importacao.total_atualizados
        )
        self.importacao.status = "SUCESSO"
        self.importacao.motivo_erro = ""
        self.importacao.save(
            update_fields=[
                "total_registros",
                "total_validos",
                "total_invalidos",
                "total_duplicados",
                "total_atualizados",
                "status",
                "motivo_erro",
            ]
        )
        # Persiste as linhas rejeitadas para a tela de resultado e o CSV de erros (HU-022).
        if self._falhas:
            FalhaImportacao.objects.bulk_create(self._falhas, batch_size=TAMANHO_LOTE)

    def _inserir_no_banco(self, df):
        if df.empty:
            return

        registros = []
        # Preencher NaNs com strings vazias para colunas de texto
        df["status_acesso"] = df["status_acesso"].fillna("")

        for _, row in df.iterrows():
            ponto_acesso_id = None
            if pd.notna(row["ponto_acesso_id"]):
                ponto_acesso_id = int(row["ponto_acesso_id"])

            registros.append(
                RegistroAcesso(
                    credencial_cifrada=row["credencial_cifrada"],
                    nome_cifrado=row["nome_cifrado"],
                    ponto_acesso_id=ponto_acesso_id,
                    tipo_acesso=row["status_acesso"],
                    timestamp=row["timestamp_dt"],
                    importacao=self.importacao,
                    evento=_texto_opcional(row.get("evento")),
                    foto=_texto_opcional(row.get("foto")),
                    tipo_consulta=_texto_opcional(row.get("tipo_consulta")),
                    area_origem=_texto_opcional(row.get("area_origem")),
                    area_destino=_texto_opcional(row.get("area_destino")),
                )
            )

        RegistroAcesso.objects.bulk_create(
            registros,
            batch_size=TAMANHO_LOTE,
        )

    def _marcar_como_falha(self, exc: Exception):
        self.importacao.status = "FALHA"
        self.importacao.motivo_erro = str(exc)[:255]
        self.importacao.save(update_fields=["status", "motivo_erro"])
